import openpyxl

def data(row_number, customers_file, quantity_file):
    wb1 = openpyxl.load_workbook(filename=customers_file)
    wb2 = openpyxl.load_workbook(filename=quantity_file)
    sheet1 = wb1['Sheet1']
    sheet2 = wb2['Sheet1']
    contact = sheet1.cell(row = row_number, column = 2).value
    email = sheet1.cell(row = row_number, column = 4).value
    phone = sheet1.cell(row = row_number, column = 5).value
    product = sheet2.cell(row = row_number-3, column = 1).value
    manufacturer = sheet2.cell(row = row_number-3, column = 2).value
    quantity = sheet2.cell(row = row_number-3, column = 3).value
    if (contact == None):
         contact = ""
    if (email == None):
         email = ""
    if (phone == None):
         phone = "+ 7 (123) 456-7890"
    if (product == None):
         product = ""
    if (manufacturer == None):
         manufacturer = ""
    if (quantity == None):
         quantity = ""
    return contact, email, phone, product, manufacturer, quantity
